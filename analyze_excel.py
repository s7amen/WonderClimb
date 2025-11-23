import json
import sys
from pathlib import Path
from collections import defaultdict
from datetime import datetime

# Fix Windows console encoding
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("openpyxl not available. Trying alternative method...")

def analyze_excel_with_openpyxl(excel_path):
    """Analyze Excel file using openpyxl"""
    
    print(f"Analyzing: {excel_path}\n")
    
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    sheets_info = {}
    
    print(f"Found {len(wb.sheetnames)} sheets:\n")
    
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
        ws = wb[sheet_name]
        
        # Get all rows
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            print(f"    (Empty sheet)")
            sheets_info[sheet_name] = {
                'name': sheet_name,
                'rows': 0,
                'columns': 0,
                'column_names': [],
                'sample_data': {}
            }
            continue
        
        # First row is likely headers
        headers = [str(cell) if cell is not None else f"Column_{i+1}" 
                  for i, cell in enumerate(rows[0])]
        
        # Remove None headers
        headers = [h for h in headers if h and h != 'None']
        
        # Analyze data rows
        data_rows = rows[1:] if len(rows) > 1 else []
        
        info = {
            'name': sheet_name,
            'rows': len(data_rows),
            'columns': len(headers),
            'column_names': headers,
            'sample_data': {},
            'data_types': {},
            'null_counts': {},
            'unique_values': {},
            'relationships': []
        }
        
        # Analyze each column
        for col_idx, col_name in enumerate(headers):
            if not col_name or col_name == 'None':
                continue
                
            column_values = []
            for row in data_rows:
                if col_idx < len(row):
                    val = row[col_idx]
                    if val is not None and str(val).strip():
                        column_values.append(val)
            
            # Data type detection
            types_found = set()
            for val in column_values[:100]:  # Sample first 100
                if isinstance(val, (int, float)):
                    types_found.add('number')
                elif isinstance(val, datetime):
                    types_found.add('date')
                elif isinstance(val, bool):
                    types_found.add('boolean')
                else:
                    types_found.add('text')
            
            info['data_types'][col_name] = list(types_found) if types_found else ['unknown']
            
            # Null counts
            info['null_counts'][col_name] = len(data_rows) - len(column_values)
            
            # Unique values (sample)
            unique_vals = list(set(str(v) for v in column_values))[:10]
            info['unique_values'][col_name] = unique_vals
            
            # Sample data
            if column_values:
                sample = str(column_values[0])
                if len(sample) > 100:
                    sample = sample[:100] + "..."
                info['sample_data'][col_name] = sample
            
            # Check for relationships
            col_lower = col_name.lower()
            if any(keyword in col_lower for keyword in ['id', 'ref', 'link', 'parent', 'child', 'climber', 'user', 'session', 'booking', 'coach', 'trainer', 'name']):
                info['relationships'].append({
                    'column': col_name,
                    'type': 'potential_reference',
                    'sample_values': [str(v) for v in column_values[:5]]
                })
        
        sheets_info[sheet_name] = info
        
        # Print summary
        print(f"    Rows: {info['rows']}, Columns: {info['columns']}")
        print(f"    Columns: {', '.join(headers[:10])}")
        if len(headers) > 10:
            print(f"    ... and {len(headers) - 10} more")
        print()
    
    wb.close()
    
    # Generate JSON report
    report = {
        'file_path': str(excel_path),
        'total_sheets': len(wb.sheetnames),
        'sheets': sheets_info,
        'analysis_date': datetime.now().isoformat()
    }
    
    # Convert to JSON-serializable format
    def convert_to_json(obj):
        if isinstance(obj, (datetime,)):
            return obj.isoformat()
        elif isinstance(obj, set):
            return list(obj)
        return obj
    
    report_serializable = json.loads(json.dumps(report, default=str))
    
    # Save report
    report_path = Path('excel_analysis_report.json')
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(report_serializable, f, indent=2, ensure_ascii=False)
    
    print(f"\n{'='*80}")
    print(f"Report saved to: {report_path}")
    print(f"{'='*80}\n")
    
    # Print detailed analysis
    print("\n" + "="*80)
    print("DETAILED SHEET ANALYSIS")
    print("="*80 + "\n")
    
    for sheet_name, info in sheets_info.items():
        print(f"\n{'='*60}")
        print(f"SHEET: {sheet_name}")
        print(f"{'='*60}")
        print(f"Rows: {info['rows']} | Columns: {info['columns']}\n")
        
        if info['column_names']:
            print("Columns and Data Types:")
            for col in info['column_names']:
                dtype = info['data_types'].get(col, ['unknown'])
                nulls = info['null_counts'].get(col, 0)
                sample = info['sample_data'].get(col, 'N/A')
                print(f"  • {col}")
                print(f"    Type: {', '.join(dtype)} | Nulls: {nulls}")
                print(f"    Sample: {sample}")
                if info['relationships']:
                    for rel in info['relationships']:
                        if rel['column'] == col:
                            print(f"    Potential reference column")
                print()
    
    return report

if __name__ == "__main__":
    excel_path = Path("old training system.xlsx")
    
    if not excel_path.exists():
        print(f"Error: File not found: {excel_path}")
        print("Please make sure the Excel file is in the project root directory.")
    else:
        if not OPENPYXL_AVAILABLE:
            print("ERROR: openpyxl library is required.")
            print("Please install it using: python -m pip install openpyxl")
            print("\nOr if pip is not available, try:")
            print("  - python -m ensurepip --upgrade")
            print("  - Then: python -m pip install openpyxl")
        else:
            try:
                report = analyze_excel_with_openpyxl(excel_path)
                print("\n[OK] Analysis complete!")
            except Exception as e:
                print(f"\n[ERROR] Error during analysis: {e}")
                import traceback
                traceback.print_exc()
